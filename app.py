import streamlit as st
import pdfplumber
import openpyxl
import subprocess
import os
import re
import datetime
from io import BytesIO

st.set_page_config(page_title="DJIM / DJIM Electrónica", page_icon="📄", layout="centered")
st.title("📄 Generador DJIM / DJIM Electrónica")

TEMPLATE_PATH = "template_djim.xlsx"

# ─────────────────────────────────────────────
# UTILIDADES PDF
# ─────────────────────────────────────────────

def extract_text_pdfplumber(pdf_bytes):
    text = ""
    try:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
    except:
        pass
    return text.strip()


def ocr_pdf_bytes(pdf_bytes, label):
    tmp_pdf = f"/tmp/{label}.pdf"
    with open(tmp_pdf, "wb") as f:
        f.write(pdf_bytes)
    subprocess.run(["pdftoppm", "-r", "300", tmp_pdf, f"/tmp/ocr_{label}"], capture_output=True)
    images = sorted([x for x in os.listdir("/tmp") if x.startswith(f"ocr_{label}")])
    text = ""
    for img in images:
        result = subprocess.run(["tesseract", f"/tmp/{img}", "stdout"], capture_output=True, text=True)
        text += result.stdout
    for img in images:
        try:
            os.remove(f"/tmp/{img}")
        except:
            pass
    return text


def get_text(pdf_bytes, label):
    text = extract_text_pdfplumber(pdf_bytes)
    if not text:
        text = ocr_pdf_bytes(pdf_bytes, label)
    return text


# ─────────────────────────────────────────────
# PARSEO DI
# ─────────────────────────────────────────────

def parsear_di(text):
    datos = {}
    alertas = []

    # Nro despacho y aduana: "25 001 IC04 258207 S"
    m = re.search(r'\b(\d{2})\s+(\d{3})\s+([A-Z]{2}\d{2})\s+(\d+)\s+([A-Z])\b', text)
    if m:
        anio, aduana, tipo, nro, dc = m.groups()
        datos['nro_despacho'] = f"{tipo}{nro}{dc}"
        datos['anio'] = anio
        datos['id_aduana'] = aduana
    else:
        alertas.append("❌ No se encontró número de despacho en el DI.")
        datos['nro_despacho'] = ''
        datos['anio'] = ''
        datos['id_aduana'] = ''

    # Fecha oficialización
    fechas = re.findall(r'\b(\d{2}/\d{2}/\d{4})\b', text)
    if fechas:
        datos['fecha_nac'] = fechas[0]
    else:
        alertas.append("❌ No se encontró fecha de oficialización en el DI.")
        datos['fecha_nac'] = ''

    # CUITs
    cuits = re.findall(r'\b(\d{2}-\d{8}-\d{1})\b', text)
    if len(cuits) >= 1:
        datos['cuit_importador'] = cuits[0]
        datos['cuit_comprador'] = cuits[0]
    else:
        alertas.append("❌ No se encontró CUIT del importador en el DI.")
        datos['cuit_importador'] = ''
        datos['cuit_comprador'] = ''

    if len(cuits) >= 2:
        datos['cuit_despachante'] = cuits[1]
    else:
        alertas.append("⚠️ No se encontró CUIT del despachante. Se usará el valor por defecto.")
        datos['cuit_despachante'] = '20-22824212-9'

    # Importador
    m = re.search(r'(FINNING\s+\S+(?:\s+\S+){1,3})', text)
    datos['importador'] = m.group(1).strip() if m else 'FINNING SOLUCIONES MINERAS SA'

    # País procedencia
    PAISES = {
        'ESTADOS UNIDOS': '212', 'JAPON': '119', 'ALEMANIA': '101',
        'BRASIL': '023', 'CHINA': '156', 'REINO UNIDO': '826',
        'SUIZA': '756', 'FRANCIA': '250', 'ITALIA': '380',
    }
    datos['pais_procedencia'] = ''
    for pais, codigo in PAISES.items():
        if pais in text.upper():
            datos['pais_procedencia'] = codigo
            break
    if not datos['pais_procedencia']:
        alertas.append("⚠️ No se encontró país de procedencia en el DI. Ingresá el código manualmente.")

    # Régimen
    m = re.search(r'REGIMEN[^0-9]*(\d{1,3})', text, re.IGNORECASE)
    datos['regimen'] = m.group(1) if m else '20'

    # Año fabricación ENGINE: ZA(XXXXXX)
    m = re.search(r'ZA\(0*(\d{4})\)', text)
    datos['anio_fab_di'] = m.group(1) if m else ''

    return datos, alertas


# ─────────────────────────────────────────────
# PARSEO DNRPA
# ─────────────────────────────────────────────

def parsear_dnrpa(text, label=""):
    datos = {}
    alertas = []

    m = re.search(r'(\d{3})\s+([A-Z]+)\s+(\w+)\s+(\w+)', text)
    if m:
        datos['id_marca'] = m.group(1)
        datos['marca_desc'] = m.group(2)
        datos['id_modelo'] = m.group(3)
        datos['cm_modelo'] = m.group(4)
    else:
        alertas.append(f"❌ No se encontró marca/modelo en DNRPA {label}.")
        datos['id_marca'] = ''
        datos['id_modelo'] = ''

    datos['tipos'] = {}
    lines = text.split('\n')
    for i, line in enumerate(lines):
        line_clean = line.strip()
        m_tipo = re.match(r'^(\d{2})\s+(BLOCK|MOTOR)', line_clean, re.IGNORECASE)
        if m_tipo:
            codigo = m_tipo.group(1)
            tipo_key = m_tipo.group(2).upper()
            contexto = line_clean + ' ' + (lines[i+1].strip() if i+1 < len(lines) else '')
            peso_m = re.search(r'(\d[\d,\.]+)\s*(KGS?|C\.C\.)', contexto, re.IGNORECASE)
            peso = peso_m.group(1).replace(',', '').replace('.', '') if peso_m else ''
            datos['tipos'][tipo_key] = {'codigo': codigo, 'peso': peso}

    if not datos['tipos']:
        alertas.append(f"❌ No se encontraron tipos (BLOCK/MOTOR) en DNRPA {label}.")

    return datos, alertas


# ─────────────────────────────────────────────
# PARSEO FACTURA
# ─────────────────────────────────────────────

def parsear_facturas(textos):
    motores = []
    texto_completo = "\n".join(textos)
    lines = texto_completo.split('\n')
    for i, line in enumerate(lines):
        if 'ENGINE' in line.upper():
            for j in range(1, 4):
                if i + j < len(lines):
                    uid = re.search(r'UNIQUE\s+ID[:\s]+(\S+)', lines[i+j], re.IGNORECASE)
                    if uid:
                        motores.append(uid.group(1))
                        break
    return motores


# ─────────────────────────────────────────────
# GENERAR TXT
# ─────────────────────────────────────────────

def generar_txt(di, items_procesados, lcm_valor):
    nro_despacho_base = di['nro_despacho']
    try:
        fecha_dt = datetime.datetime.strptime(di['fecha_nac'], "%d/%m/%Y")
        anio_dos = str(fecha_dt.year)[-2:]
        fecha_str = fecha_dt.strftime("%d/%m/%Y")
    except:
        anio_dos = di.get('anio', '26')
        fecha_str = di.get('fecha_nac', '')

    nro_despacho = f"{nro_despacho_base}/{anio_dos}"
    id_aduana = di.get('id_aduana', '001')
    cantidad_lineas = str(len(items_procesados))

    if lcm_valor and lcm_valor.strip():
        parts = (re.split(r'[/\-\s]+', lcm_valor.strip()) + ["0", "0", "0"])[:3]
        lcm_tipo, lcm_nro, lcm_anio = parts
    else:
        lcm_tipo, lcm_nro, lcm_anio = "0", "0", "0"

    def q(v): return f'"{v}"'
    def safe(v): return str(v).strip().replace(" ", "") if v else ""

    caratula = ";".join([
        q(id_aduana), q(nro_despacho), q("00"), q("12"),
        q(di.get('cuit_importador', '')), q("12"),
        q(di.get('cuit_comprador', '')), q("12"),
        q(di.get('cuit_despachante', '')), q(di.get('regimen', '20')),
        q(fecha_str), q(di.get('pais_procedencia', '212')),
        q(cantidad_lineas), q("N"), q("S"),
        q(""), q(""), q(""), q(""), q("")
    ])

    lineas = []
    for i, item in enumerate(items_procesados, start=1):
        dnrpa = item['dnrpa']
        tipo = item['tipo']
        anio = item['anio_fab']
        nro_motor = safe(item.get('motor', '')) if tipo == 'ENGINE' else ''
        tipo_key = 'MOTOR' if tipo == 'ENGINE' else 'BLOCK'
        id_tipo = dnrpa.get('tipos', {}).get(tipo_key, {}).get('codigo', '')
        peso = dnrpa.get('tipos', {}).get(tipo_key, {}).get('peso', '')

        linea = ";".join([
            q(id_aduana), q(nro_despacho), q("00"), q(str(i)),
            q(dnrpa.get('id_marca', '')), q(id_tipo),
            q(dnrpa.get('id_modelo', '')),
            q(lcm_tipo), q(lcm_nro), q(lcm_anio),
            q(str(anio)), q(str(anio)),
            q(dnrpa.get('id_marca', '')), q(nro_motor),
            q("000"), q("NOPOSEE"),
            q(di.get('pais_procedencia', '212')), q(str(peso)), q("N")
        ])
        lineas.append(linea)

    return caratula + "\n" + "\n".join(lineas)


# ─────────────────────────────────────────────
# GENERAR EXCEL
# ─────────────────────────────────────────────

def generar_excel(di, items_procesados, lcm_valor):
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb['ANVERSO']

    try:
        fecha_dt = datetime.datetime.strptime(di['fecha_nac'], "%d/%m/%Y")
    except:
        fecha_dt = datetime.datetime.now()

    ws['E3'] = di['nro_despacho']
    ws['J3'] = fecha_dt
    ws['L3'] = di.get('regimen', '20')
    ws['E7'] = di.get('importador', '')
    ws['L7'] = di.get('cuit_importador', '')
    ws['I9'] = di.get('importador', '')
    ws['L9'] = di.get('cuit_comprador', '')
    try:
        ws['E11'] = int(di.get('pais_procedencia', 212))
    except:
        ws['E11'] = di.get('pais_procedencia', 212)

    for row_idx in range(16, 31):
        for col_idx in range(1, 14):
            ws.cell(row=row_idx, column=col_idx).value = None

    lcm_excel = lcm_valor.strip() if lcm_valor and lcm_valor.strip() else 'XXX'

    for i, item in enumerate(items_procesados):
        row = 16 + i
        dnrpa = item['dnrpa']
        tipo = item['tipo']
        anio = item['anio_fab']
        nro_motor = item.get('motor', '') if tipo == 'ENGINE' else ''
        tipo_key = 'MOTOR' if tipo == 'ENGINE' else 'BLOCK'
        id_tipo = dnrpa.get('tipos', {}).get(tipo_key, {}).get('codigo', '')
        peso = dnrpa.get('tipos', {}).get(tipo_key, {}).get('peso', '')

        ws.cell(row=row, column=1).value = i + 1
        ws.cell(row=row, column=2).value = dnrpa.get('id_marca', '')
        ws.cell(row=row, column=3).value = id_tipo
        ws.cell(row=row, column=4).value = dnrpa.get('id_modelo', '')
        ws.cell(row=row, column=5).value = lcm_excel
        ws.cell(row=row, column=6).value = str(anio)
        ws.cell(row=row, column=7).value = str(anio)
        ws.cell(row=row, column=8).value = dnrpa.get('id_marca', '')
        ws.cell(row=row, column=9).value = nro_motor
        ws.cell(row=row, column=10).value = '000'
        ws.cell(row=row, column=11).value = 'NO POSEE'
        ws.cell(row=row, column=12).value = di.get('pais_procedencia', '212')
        ws.cell(row=row, column=13).value = str(peso)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
# INTERFAZ
# ─────────────────────────────────────────────

if 'items' not in st.session_state:
    st.session_state.items = []

st.markdown("### 1. Documentos generales")
col1, col2 = st.columns(2)
with col1:
    di_file = st.file_uploader("📋 DI (PDF)", type="pdf")
with col2:
    fc_files = st.file_uploader("🧾 Factura/s (PDF)", type="pdf", accept_multiple_files=True)

st.markdown("### 2. Ítems de la DJIM")
st.caption("Agregá un ítem por cada motor o block del despacho.")

if st.button("➕ Agregar ítem"):
    st.session_state.items.append({'tipo': 'ENGINE', 'dnrpa_file': None, 'anio_fab_manual': ''})

items_to_remove = []
for idx, item in enumerate(st.session_state.items):
    with st.container(border=True):
        col1, col2, col3 = st.columns([2, 3, 1])
        with col1:
            tipo = st.selectbox(
                f"Tipo ítem {idx+1}", ["ENGINE", "BLOCK"],
                key=f"tipo_{idx}",
                index=0 if item['tipo'] == 'ENGINE' else 1
            )
            st.session_state.items[idx]['tipo'] = tipo
        with col2:
            dnrpa = st.file_uploader(f"DNRPA PDF - ítem {idx+1}", type="pdf", key=f"dnrpa_{idx}")
            st.session_state.items[idx]['dnrpa_file'] = dnrpa
        with col3:
            st.markdown("<br><br>", unsafe_allow_html=True)
            if st.button("🗑️", key=f"remove_{idx}"):
                items_to_remove.append(idx)
        if tipo == 'BLOCK':
            anio = st.text_input(
                f"Año fabricación ítem {idx+1}",
                value=item.get('anio_fab_manual', ''),
                key=f"anio_{idx}"
            )
            st.session_state.items[idx]['anio_fab_manual'] = anio

for idx in sorted(items_to_remove, reverse=True):
    st.session_state.items.pop(idx)

st.markdown("### 3. Datos adicionales")
col1, col2 = st.columns(2)
with col1:
    pais_fab_manual = st.text_input("Código país fabricación", value="212",
                                     help="Solo si no se detecta automáticamente del DI")
    tiene_lcm = st.radio("¿Tiene LCM?", ["No", "Sí"], horizontal=True)
with col2:
    lcm_valor = ""
    if tiene_lcm == "Sí":
        lcm_valor = st.text_input("Número LCM", placeholder="ej: 39/12345/2025")

st.markdown("---")

if st.button("⚙️ Procesar y Generar", type="primary", use_container_width=True):

    errores = []
    if not di_file:
        errores.append("❌ Faltá subir el DI.")
    if not fc_files:
        errores.append("❌ Faltá subir al menos una factura.")
    if not st.session_state.items:
        errores.append("❌ Agregá al menos un ítem.")
    for idx, item in enumerate(st.session_state.items):
        if not item.get('dnrpa_file'):
            errores.append(f"❌ Faltá el DNRPA del ítem {idx+1}.")
        if item['tipo'] == 'BLOCK' and not item.get('anio_fab_manual', '').strip():
            errores.append(f"❌ Ingresá el año de fabricación del ítem {idx+1} (BLOCK).")

    if errores:
        for e in errores:
            st.error(e)
        st.stop()

    with st.spinner("Procesando documentos..."):
        di_bytes = di_file.read()
        di_text = get_text(di_bytes, "di")
        di_datos, di_alertas = parsear_di(di_text)
        if not di_datos['pais_procedencia']:
            di_datos['pais_procedencia'] = pais_fab_manual

        fc_textos = []
        for i, fc_f in enumerate(fc_files):
            t = get_text(fc_f.read(), f"fc_{i}")
            fc_textos.append(t)
        motores_factura = parsear_facturas(fc_textos)

        n_engines = sum(1 for it in st.session_state.items if it['tipo'] == 'ENGINE')
        items_procesados = []
        todas_alertas = di_alertas.copy()
        motor_idx = 0

        for idx, item in enumerate(st.session_state.items):
            dnrpa_bytes = item['dnrpa_file'].read()
            dnrpa_text = get_text(dnrpa_bytes, f"dnrpa_{idx}")
            dnrpa_datos, dnrpa_alertas = parsear_dnrpa(dnrpa_text, label=f"ítem {idx+1}")
            todas_alertas.extend(dnrpa_alertas)

            tipo = item['tipo']
            tipo_key = 'MOTOR' if tipo == 'ENGINE' else 'BLOCK'

            if tipo == 'ENGINE':
                anio_fab = di_datos.get('anio_fab_di', '')
                if not anio_fab:
                    todas_alertas.append(f"❌ No se encontró año de fabricación en el DI para ENGINE ítem {idx+1}.")
            else:
                anio_fab = item.get('anio_fab_manual', '')

            motor = ''
            if tipo == 'ENGINE':
                if motor_idx < len(motores_factura):
                    motor = motores_factura[motor_idx]
                    motor_idx += 1
                else:
                    todas_alertas.append(f"❌ No se encontró UNIQUE ID para ENGINE ítem {idx+1} en la factura.")

            if not dnrpa_datos.get('tipos', {}).get(tipo_key, {}).get('peso'):
                todas_alertas.append(f"❌ No se encontró peso para {tipo} en DNRPA ítem {idx+1}.")

            items_procesados.append({
                'tipo': tipo,
                'dnrpa': dnrpa_datos,
                'anio_fab': anio_fab,
                'motor': motor,
            })

        if n_engines > len(motores_factura):
            todas_alertas.append(
                f"❌ Se declararon {n_engines} ENGINE(s) pero se encontraron "
                f"solo {len(motores_factura)} UNIQUE ID(s) en la/s factura/s."
            )

    advertencias = [a for a in todas_alertas if a.startswith("⚠️")]
    errores_criticos = [a for a in todas_alertas if a.startswith("❌")]

    for a in advertencias:
        st.warning(a)

    if errores_criticos:
        for e in errores_criticos:
            st.error(e)
        st.stop()

    st.success("✅ Documentos procesados correctamente.")

    with st.expander("📋 Ver datos extraídos"):
        st.markdown("**DI:**")
        st.json({k: v for k, v in di_datos.items() if k != 'anio_fab_di'})
        for idx, item in enumerate(items_procesados):
            st.markdown(f"**Ítem {idx+1} ({item['tipo']}):**")
            st.json({
                'id_marca': item['dnrpa'].get('id_marca'),
                'id_modelo': item['dnrpa'].get('id_modelo'),
                'tipos': item['dnrpa'].get('tipos'),
                'anio_fab': item['anio_fab'],
                'motor': item.get('motor', ''),
            })

    st.markdown("### 4. Descargar")

    txt_content = generar_txt(di_datos, items_procesados, lcm_valor)
    st.download_button(
        "📥 Descargar DJIM Electrónica (.txt)",
        data=txt_content.encode('utf-8'),
        file_name="DJIM_ELECTRONICA.txt",
        mime="text/plain",
        use_container_width=True
    )

    if os.path.exists(TEMPLATE_PATH):
        excel_buf = generar_excel(di_datos, items_procesados, lcm_valor)
        nro = di_datos.get('nro_despacho', 'DJIM')
        st.download_button(
            "📥 Descargar DJIM Excel (.xlsx)",
            data=excel_buf,
            file_name=f"DJIM_{nro}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.warning("⚠️ Template Excel no encontrado en el repo.")
